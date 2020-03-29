from openpyxl import load_workbook


def get_rows_count(file_path, sheet_name):
    input_wb = load_workbook(file_path)
    input_sheet = input_wb[sheet_name]

    return input_sheet.max_row


def get_columns_count(file_path, sheet_name):
    input_wb = load_workbook(file_path)
    input_sheet = input_wb[sheet_name]

    return input_sheet.max_column


def read_cell_data(fille_path, sheet_name, row_num, column_num):
    input_wb = load_workbook(fille_path)
    input_sheet = input_wb[sheet_name]

    return input_sheet.cell(row_num, column_num).value


def write_data_into_cell(file_path, sheet_name, row_num, column_num, data):
    input_wb = load_workbook(file_path)
    input_sheet = input_wb[sheet_name]
    input_sheet.cell(row_num, column_num).value = data
    input_wb.save(file_path)
