from openpyxl import load_workbook

path = r"C:\Users\Umang Bhatia\Documents\ReadExcel.xlsx"

wb = load_workbook(path)

sheet = wb['Sheet1']

rows = sheet.max_row
columns = sheet.max_column

print(rows)
print(columns)

for r in range(1, rows + 1):
    for c in range(1, columns + 1):
        print(sheet.cell(r, c).value, end=' | ')
    print('\n')
