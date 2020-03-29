from openpyxl import load_workbook

excel_path = r"C:\Users\Umang Bhatia\Documents\WriteExcel.xlsx"

input_wb = load_workbook(excel_path)

sheet = input_wb['Sheet1']

for r in range(1, 5):
    for c in range(1, 4):
        sheet.cell(r, c).value = 'welcome'

input_wb.save(excel_path)
